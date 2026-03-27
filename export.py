"""
SGAM - Sistema de Gestión de Asistencias Médicas
Módulo: export.py
Responsabilidad: Generación de reportes Excel institucionales con openpyxl.

Mejoras v1.2:
  - Reporte consolidado en HOJA ÚNICA compacta: una fila por médico,
    columnas de días coloreadas (optimizado para impresión).
  - Exportación individual por empleado (sin cambios de estructura).
  - Filtros por Tipo y Especialidad aplicables a ambos modos.
"""

import calendar
from datetime import date
from pathlib import Path
from typing import Optional, Callable

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from core import COLOR_MAP


# ──────────────────────────────────────────────
# Estilos reutilizables
# ──────────────────────────────────────────────
def _font(bold=False, size=9, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


FILL_NAVY    = _fill("1F3864")
FILL_BLUE    = _fill("2E75B6")
FILL_HEADER  = _fill("1F3864")
FILL_LBLUE   = _fill("DEEAF1")
FILL_XLIGHT  = _fill("EBF3FB")
FILL_WHITE   = _fill("FFFFFF")

FONT_TITLE   = _font(bold=True, size=14, color="1F3864")
FONT_H_WHITE = _font(bold=True, size=9, color="FFFFFF")
FONT_H_NAVY  = _font(bold=True, size=9, color="1F3864")
FONT_BODY    = _font(size=9)
FONT_SMALL   = _font(size=7)
FONT_ITALIC  = _font(size=9, italic=True, color="595959")

ALIN_C  = _align("center", "center")
ALIN_L  = _align("left",   "center")
ALIN_CW = _align("center", "center", wrap=True)
BORDER  = _border("thin")
BORDER_M = _border("medium")

MESES_ES = {
    1: "Enero",    2: "Febrero",  3: "Marzo",     4: "Abril",
    5: "Mayo",     6: "Junio",    7: "Julio",      8: "Agosto",
    9: "Sept.",   10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}
DIAS_ES = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sá", "Do"]


# ──────────────────────────────────────────────
# Helpers internos
# ──────────────────────────────────────────────
def _cell(ws, row, col, value="", fill=None, font=None,
          align=ALIN_C, border=BORDER):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill   = fill
    if font:   c.font   = font
    c.alignment = align
    c.border    = border
    return c


def _insertar_logo(ws, ruta_logo: Optional[str], celda: str = "A1"):
    if not ruta_logo:
        return
    ruta = Path(ruta_logo)
    if not ruta.exists():
        return
    try:
        img = XLImage(str(ruta))
        img.width  = 160
        img.height = 60
        ws.add_image(img, celda)
    except Exception:
        pass


def _nombre_archivo(nombre: str, mes: int, anio: int) -> str:
    limpio = nombre.replace(" ", "_").replace(",", "")
    return f"SGAM_Reporte_{limpio}_{MESES_ES.get(mes, mes)}{anio}.xlsx"


def _aplicar_filtros(df: pd.DataFrame, filtro_tipo: Optional[str],
                      filtro_especialidad: Optional[str]) -> pd.DataFrame:
    """Aplica filtros opcionales por Tipo y Especialidad."""
    df_f = df.copy()
    if filtro_tipo and filtro_tipo.strip():
        df_f = df_f[df_f["Tipo"].str.lower().str.contains(
            filtro_tipo.strip().lower(), na=False)]
    if filtro_especialidad and filtro_especialidad.strip():
        df_f = df_f[df_f["Especialidad_Base"].str.lower().str.contains(
            filtro_especialidad.strip().lower(), na=False)]
    if df_f.empty:
        raise ValueError(
            f"Ningún empleado coincide con los filtros:\n"
            f"  Tipo: '{filtro_tipo or 'Todos'}'\n"
            f"  Especialidad: '{filtro_especialidad or 'Todas'}'"
        )
    return df_f


# ──────────────────────────────────────────────
# Reporte Individual (hoja de calendario)
# ──────────────────────────────────────────────
def _encabezado_individual(ws, empleado: dict, mes: int, anio: int,
                             ruta_logo: Optional[str]):
    """Bloque de encabezado institucional para reporte individual."""
    ws.row_dimensions[1].height = 50
    ws.merge_cells("A1:C1")
    _insertar_logo(ws, ruta_logo, "A1")

    ws.merge_cells("D1:K1")
    c = ws.cell(row=1, column=4,
                 value="SGAM – Sistema de Gestión de Asistencias Médicas")
    c.font = FONT_TITLE; c.alignment = ALIN_C; c.fill = FILL_XLIGHT

    ws.merge_cells("A2:K2")
    c2 = ws.cell(row=2, column=1,
                  value="HOSPITAL GENERAL – Departamento de Recursos Humanos")
    c2.font = _font(bold=True, size=11, color="1F3864")
    c2.alignment = ALIN_C; c2.fill = FILL_LBLUE

    ws.row_dimensions[3].height = 6

    campos = [
        ("Nombre:",        empleado.get("Nombre_Completo", "")),
        ("ID Biométrico:", empleado.get("ID_Institucional", "")),
        ("Tipo:",          empleado.get("Tipo", "")),
        ("Especialidad:",  empleado.get("Especialidad_Base", "")),
        ("Período:",       f"{MESES_ES.get(mes, mes)} {anio}"),
    ]
    for i, (etq, val) in enumerate(campos):
        fila = 4 + i
        ws.row_dimensions[fila].height = 16
        c_et = ws.cell(row=fila, column=1, value=etq)
        c_et.font = FONT_H_NAVY; c_et.alignment = ALIN_L
        ws.merge_cells(f"B{fila}:E{fila}")
        c_val = ws.cell(row=fila, column=2, value=val)
        c_val.font = FONT_BODY; c_val.alignment = ALIN_L
        c_val.border = Border(bottom=Side(style="thin"))


def _calendario_individual(ws, df_emp: pd.DataFrame,
                              mes: int, anio: int, fila_ini: int = 11) -> int:
    """Calendario mensual coloreado en una hoja individual."""
    dias_mes = calendar.monthrange(anio, mes)[1]

    # Título del calendario
    ws.row_dimensions[fila_ini].height = 20
    ws.merge_cells(
        start_row=fila_ini, start_column=1,
        end_row=fila_ini,   end_column=dias_mes + 1
    )
    c = ws.cell(row=fila_ini, column=1,
                 value=f"CALENDARIO – {MESES_ES.get(mes, mes).upper()} {anio}")
    c.font = FONT_H_WHITE; c.fill = FILL_NAVY; c.alignment = ALIN_C

    # Fila de números de día
    fila_d = fila_ini + 1
    ws.row_dimensions[fila_d].height = 18
    _cell(ws, fila_d, 1, "Día →", fill=FILL_BLUE, font=FONT_H_WHITE)
    for d in range(1, dias_mes + 1):
        col = d + 1
        ws.column_dimensions[get_column_letter(col)].width = 5
        _cell(ws, fila_d, col, d, fill=FILL_BLUE, font=FONT_H_WHITE)
        fecha_d = date(anio, mes, d)
        c_ds = ws.cell(row=fila_d + 1, column=col,
                        value=DIAS_ES[fecha_d.weekday()])
        c_ds.font = _font(size=7, color="404040")
        c_ds.alignment = ALIN_C; c_ds.border = BORDER

    # Fila de estatus coloreado
    fila_st = fila_d + 2
    ws.row_dimensions[fila_st].height = 22
    _cell(ws, fila_st, 1, "Estatus", fill=FILL_BLUE, font=FONT_H_WHITE)

    mapa_dia = {}
    for _, row in df_emp.iterrows():
        d = pd.to_datetime(row["Fecha"]).day
        mapa_dia[d] = {
            "label": row.get("Label", ""),
            "color": row.get("Color_Hex", "FFFFFF"),
            "notas": row.get("Notas", ""),
        }

    for d in range(1, dias_mes + 1):
        col  = d + 1
        info = mapa_dia.get(d, {"label": "", "color": "FFFFFF", "notas": ""})
        c = ws.cell(row=fila_st, column=col, value=(info["label"] or "")[:3])
        c.fill = _fill(info["color"]); c.font = FONT_SMALL
        c.alignment = ALIN_C; c.border = BORDER
        if info["notas"]:
            from openpyxl.comments import Comment
            cm = Comment(info["notas"], "SGAM")
            cm.width = 150; cm.height = 60
            c.comment = cm

    # Leyenda
    fila_ley = fila_st + 2
    ws.merge_cells(
        start_row=fila_ley, start_column=1,
        end_row=fila_ley,   end_column=dias_mes + 1
    )
    c_ley = ws.cell(row=fila_ley, column=1,
                     value="LEYENDA: 🟡 Asistencia  🔴 Retardo  🟠 Falta  "
                           "⚪ No Laborable  🟢 Vacaciones  🔵 Incapacidad")
    c_ley.font = _font(size=8, italic=True, color="404040")
    c_ley.alignment = ALIN_L
    return fila_ley + 2


def _seccion_firma(ws, fila_ini: int, dias_mes: int):
    ws.row_dimensions[fila_ini].height = 40
    ws.row_dimensions[fila_ini + 1].height = 16
    ws.row_dimensions[fila_ini + 2].height = 16
    ws.merge_cells(
        start_row=fila_ini, start_column=1,
        end_row=fila_ini,   end_column=10
    )
    fl = fila_ini + 1
    ws.merge_cells(f"A{fl}:E{fl}")
    ws.cell(row=fl, column=1).border = Border(bottom=Side(style="medium"))
    ws.merge_cells(f"G{fl}:K{fl}")
    ws.cell(row=fl, column=7).border = Border(bottom=Side(style="medium"))
    fe = fila_ini + 2
    ws.merge_cells(f"A{fe}:E{fe}")
    c_rh = ws.cell(row=fe, column=1, value="Firma y Sello – Recursos Humanos")
    c_rh.font = FONT_ITALIC; c_rh.alignment = ALIN_C
    ws.merge_cells(f"G{fe}:K{fe}")
    c_med = ws.cell(row=fe, column=7, value="Firma – Médico / Residente")
    c_med.font = FONT_ITALIC; c_med.alignment = ALIN_C


def exportar_reporte_empleado(df_empleado: pd.DataFrame,
                               directorio_salida: str,
                               ruta_logo: Optional[str] = None) -> str:
    """Genera reporte Excel individual de un empleado."""
    if df_empleado.empty:
        raise ValueError("DataFrame vacío.")

    primera = df_empleado.iloc[0]
    empleado = {
        "Nombre_Completo":   str(primera.get("Nombre_Completo", "Desconocido")),
        "ID_Institucional":  str(primera.get("ID_Institucional", "")),
        "Tipo":              str(primera.get("Tipo", "")),
        "Especialidad_Base": str(primera.get("Especialidad_Base", "")),
    }

    fechas   = pd.to_datetime(df_empleado["Fecha"])
    mes      = int(fechas.dt.month.mode()[0])
    anio     = int(fechas.dt.year.mode()[0])
    dias_mes = calendar.monthrange(anio, mes)[1]

    df_mes = df_empleado[
        (fechas.dt.month == mes) & (fechas.dt.year == anio)
    ].copy()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MESES_ES.get(mes, mes)} {anio}"
    ws.column_dimensions["A"].width = 14

    _encabezado_individual(ws, empleado, mes, anio, ruta_logo)
    fila_sig = _calendario_individual(ws, df_mes, mes, anio, fila_ini=11)
    _seccion_firma(ws, fila_ini=fila_sig, dias_mes=dias_mes)
    ws.freeze_panes = "A12"

    nombre = _nombre_archivo(empleado["Nombre_Completo"], mes, anio)
    Path(directorio_salida).mkdir(parents=True, exist_ok=True)
    ruta_out = Path(directorio_salida) / nombre
    wb.save(str(ruta_out))
    return str(ruta_out)


def exportar_todos(df_resultado: pd.DataFrame,
                    directorio_salida: str,
                    ruta_logo: Optional[str] = None,
                    callback: Optional[Callable] = None) -> list[str]:
    """Exporta un archivo individual por cada empleado."""
    ids    = df_resultado["ID_Institucional"].unique()
    total  = len(ids)
    rutas  = []

    for i, id_emp in enumerate(ids):
        df_emp = df_resultado[df_resultado["ID_Institucional"] == id_emp].copy()
        nombre = str(df_emp.iloc[0].get("Nombre_Completo", id_emp))
        try:
            ruta = exportar_reporte_empleado(df_emp, directorio_salida, ruta_logo)
            rutas.append(ruta)
        except Exception as e:
            print(f"[ERROR] {nombre}: {e}")
        if callback:
            callback(progreso=(i + 1) / total, nombre=nombre)

    return rutas


def exportar_filtrado(df_resultado: pd.DataFrame,
                       directorio_salida: str,
                       filtro_tipo: Optional[str] = None,
                       filtro_especialidad: Optional[str] = None,
                       ruta_logo: Optional[str] = None,
                       callback: Optional[Callable] = None) -> list[str]:
    """Exporta archivos individuales con filtros opcionales."""
    df_f = _aplicar_filtros(df_resultado, filtro_tipo, filtro_especialidad)
    return exportar_todos(df_f, directorio_salida, ruta_logo, callback)


# ──────────────────────────────────────────────
# Reporte Maestro — HOJA ÚNICA COMPACTA
# ──────────────────────────────────────────────
def exportar_maestro_consolidado(df_resultado: pd.DataFrame,
                                  directorio_salida: str,
                                  ruta_logo: Optional[str] = None,
                                  filtro_tipo: Optional[str] = None,
                                  filtro_especialidad: Optional[str] = None,
                                  callback: Optional[Callable] = None) -> str:
    """
    Genera un único archivo Excel con HOJA ÚNICA compacta.

    Diseño: una fila por médico + columnas de días (1–31) coloreadas.
    Columnas fijas: # | Nombre | Tipo | Especialidad | Grado | Periodo | d1 … d31
    Optimizado para impresión: altura de fila mínima, fuente pequeña,
    sin celdas combinadas en el área de datos.
    """
    df_f = _aplicar_filtros(df_resultado, filtro_tipo, filtro_especialidad)

    fechas   = pd.to_datetime(df_f["Fecha"])
    mes      = int(fechas.dt.month.mode()[0])
    anio     = int(fechas.dt.year.mode()[0])
    dias_mes = calendar.monthrange(anio, mes)[1]
    mes_str  = MESES_ES.get(mes, str(mes))

    # Nombre de archivo
    sfx = ""
    if filtro_tipo:
        sfx += f"_{filtro_tipo.replace(' ', '_')}"
    if filtro_especialidad:
        sfx += f"_{filtro_especialidad.replace(' ', '_')}"
    nombre_arch = f"SGAM_Maestro{sfx}_{mes_str}{anio}.xlsx"

    Path(directorio_salida).mkdir(parents=True, exist_ok=True)
    ruta_out = Path(directorio_salida) / nombre_arch

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Asistencias {mes_str} {anio}"

    # ── Fila 1: Logo + Título ────────────────────────────────────────────
    ws.row_dimensions[1].height = 46
    ws.merge_cells("A1:B1")
    _insertar_logo(ws, ruta_logo, "A1")

    # Calcular última columna del reporte (columnas fijas + días)
    N_COLS_FIJAS = 6   # #, Nombre, Tipo, Especialidad, Grado, Periodo
    ULTIMA_COL   = N_COLS_FIJAS + dias_mes
    ultimo_letra = get_column_letter(ULTIMA_COL)

    ws.merge_cells(f"C1:{ultimo_letra}1")
    c_tit = ws.cell(row=1, column=3,
                     value=f"SGAM – Reporte Maestro de Asistencias  |  "
                           f"{mes_str} {anio}"
                           + (f"  |  {filtro_tipo}" if filtro_tipo else "")
                           + (f"  |  {filtro_especialidad}" if filtro_especialidad else ""))
    c_tit.font      = _font(bold=True, size=12, color="1F3864")
    c_tit.fill      = FILL_XLIGHT
    c_tit.alignment = ALIN_L

    # ── Fila 2: Sub-encabezado hospital ─────────────────────────────────
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f"A2:{ultimo_letra}2")
    c_h = ws.cell(row=2, column=1,
                   value="HOSPITAL GENERAL – Departamento de Recursos Humanos")
    c_h.font = _font(bold=True, size=10, color="1F3864")
    c_h.fill = FILL_LBLUE; c_h.alignment = ALIN_C

    # ── Fila 3: Encabezado de columnas ──────────────────────────────────
    FILA_HEADER = 3
    ws.row_dimensions[FILA_HEADER].height = 22

    # Columnas fijas
    cols_fijas = ["#", "Nombre Completo", "Tipo", "Especialidad", "Grado", "Periodo"]
    anchos_fijos = [4, 28, 13, 16, 6, 8]

    for j, (etq, ancho) in enumerate(zip(cols_fijas, anchos_fijos), start=1):
        c = ws.cell(row=FILA_HEADER, column=j, value=etq)
        c.font = FONT_H_WHITE; c.fill = FILL_NAVY
        c.alignment = ALIN_CW; c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = ancho

    # Columnas de días
    for d in range(1, dias_mes + 1):
        col = N_COLS_FIJAS + d
        fecha_d = date(anio, mes, d)
        # Encabezado: número + día semana en 2 líneas
        c = ws.cell(row=FILA_HEADER, column=col,
                     value=f"{d}\n{DIAS_ES[fecha_d.weekday()]}")
        c.font      = _font(bold=True, size=8, color="FFFFFF")
        c.fill      = FILL_BLUE
        c.alignment = ALIN_CW
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = 3.8

    # ── Filas de empleados ───────────────────────────────────────────────
    ids    = df_f["ID_Institucional"].unique()
    total  = len(ids)

    for num_idx, id_emp in enumerate(ids, start=1):
        fila = FILA_HEADER + num_idx
        ws.row_dimensions[fila].height = 14

        df_emp  = df_f[df_f["ID_Institucional"] == id_emp]
        primera = df_emp.iloc[0]

        nombre = str(primera.get("Nombre_Completo", id_emp))
        tipo_e = str(primera.get("Tipo", ""))
        esp_e  = str(primera.get("Especialidad_Base", ""))
        grado  = str(primera.get("Grado", ""))
        period = str(primera.get("Periodo_Ingreso", ""))

        # Alternar fondo de fila para facilitar lectura
        fill_fila = _fill("F2F7FC") if num_idx % 2 == 0 else FILL_WHITE

        vals_fijos = [num_idx, nombre, tipo_e, esp_e, grado, period]
        for j, val in enumerate(vals_fijos, start=1):
            c = ws.cell(row=fila, column=j, value=val)
            c.font      = _font(size=8)
            c.fill      = fill_fila
            c.alignment = ALIN_L if j == 2 else ALIN_C
            c.border    = BORDER

        # Construir mapa {dia: info} para el empleado
        mapa_dia: dict[int, dict] = {}
        df_mes_emp = df_emp[
            (pd.to_datetime(df_emp["Fecha"]).dt.month == mes) &
            (pd.to_datetime(df_emp["Fecha"]).dt.year  == anio)
        ]
        for _, row in df_mes_emp.iterrows():
            d_num = pd.to_datetime(row["Fecha"]).day
            mapa_dia[d_num] = {
                "color": row.get("Color_Hex", "FFFFFF"),
                "label": (row.get("Label", "") or "")[:2],
                "notas": row.get("Notas", ""),
            }

        # Celdas de días coloreadas
        for d in range(1, dias_mes + 1):
            col  = N_COLS_FIJAS + d
            info = mapa_dia.get(d, {"color": "FFFFFF", "label": "", "notas": ""})
            c = ws.cell(row=fila, column=col, value=info["label"])
            c.fill      = _fill(info["color"])
            c.font      = _font(size=7, bold=True)
            c.alignment = ALIN_C
            c.border    = BORDER
            if info["notas"]:
                from openpyxl.comments import Comment
                cm = Comment(info["notas"], "SGAM")
                cm.width = 140; cm.height = 50
                c.comment = cm

        if callback:
            callback(progreso=num_idx / total, nombre=nombre)

    # ── Fila de leyenda ──────────────────────────────────────────────────
    fila_ley = FILA_HEADER + total + 2
    ws.row_dimensions[fila_ley].height = 13
    ws.merge_cells(f"A{fila_ley}:{ultimo_letra}{fila_ley}")
    c_ley = ws.cell(
        row=fila_ley, column=1,
        value="LEYENDA: 🟡As=Asistencia  🔴Re=Retardo  🟠Fa=Falta  "
              "⚪NL=No Laborable  🟢Va=Vacaciones  🔵In=Incapacidad  "
              "Pe=Permiso  Co=Comisión"
    )
    c_ley.font = _font(size=8, italic=True, color="404040")
    c_ley.alignment = ALIN_L

    # ── Configuración de página para impresión ───────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale   = 80

    # Congelar columnas fijas al desplazar horizontalmente
    ws.freeze_panes = f"G{FILA_HEADER + 1}"

    wb.save(str(ruta_out))
    return str(ruta_out)
