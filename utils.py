"""
SGAM - Sistema de Gestión de Asistencias Médicas
Módulo: utils.py
Responsabilidad: Estadísticas avanzadas, rankings y utilidades transversales.

Funciones disponibles:
  - estadisticas_por_empleado()   → DataFrame con métricas individuales
  - ranking_asistencia()          → Ordenados por % asistencia
  - resumen_por_tipo()            → Agregado por Tipo (Interno/Residente/Adscrito)
  - resumen_por_especialidad()    → Agregado por Especialidad
  - dias_criticos()               → Días con más faltas o retardos en el período
  - calcular_tendencia_semanal()  → Distribución de incidencias por semana
  - exportar_estadisticas_excel() → Hoja de estadísticas en Excel
"""

from __future__ import annotations

import calendar
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# Constantes internas
# ──────────────────────────────────────────────
ESTATUSES_LABORABLES = {"ASISTENCIA", "RETARDO", "FALTA", "VACACIONES",
                         "INCAPACIDAD", "PERMISO", "COMISION", "ROTACION", "OTRO"}
ESTATUSES_PRESENCIA  = {"ASISTENCIA", "RETARDO"}   # físicamente en el hospital

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo",
    6: "Junio", 7: "Julio", 8: "Agosto", 9: "Septiembre",
    10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


# ──────────────────────────────────────────────
# Helpers privados
# ──────────────────────────────────────────────
def _safe_pct(numerador: int, denominador: int) -> float:
    return round((numerador / denominador * 100), 1) if denominador > 0 else 0.0


def _conteos_empleado(df_emp: pd.DataFrame) -> dict:
    """Cuenta estatus para un DataFrame de un solo empleado."""
    lab = df_emp[df_emp["Estatus"].isin(ESTATUSES_LABORABLES)]
    c   = lab["Estatus"].value_counts().to_dict()
    total = len(lab)
    return {
        "total_lab":    total,
        "asistencias":  c.get("ASISTENCIA",  0),
        "retardos":     c.get("RETARDO",     0),
        "faltas":       c.get("FALTA",       0),
        "vacaciones":   c.get("VACACIONES",  0),
        "incapacidad":  c.get("INCAPACIDAD", 0),
        "permisos":     c.get("PERMISO",     0),
        "comisiones":   c.get("COMISION",    0),
        "rotaciones":   c.get("ROTACION",    0),
        "otros":        c.get("OTRO",        0),
        "pct_asistencia": _safe_pct(c.get("ASISTENCIA", 0), total),
        "pct_retardo":    _safe_pct(c.get("RETARDO",    0), total),
        "pct_falta":      _safe_pct(c.get("FALTA",      0), total),
        "pct_presencia":  _safe_pct(
            c.get("ASISTENCIA", 0) + c.get("RETARDO", 0), total
        ),
    }


# ──────────────────────────────────────────────
# Estadísticas individuales
# ──────────────────────────────────────────────
def estadisticas_por_empleado(df: pd.DataFrame) -> pd.DataFrame:
    """
    Genera un DataFrame con una fila por empleado y sus métricas del período.

    Columnas resultantes:
        ID, Nombre_Completo, Tipo, Especialidad, Grado,
        Periodo_Ingreso, total_lab, asistencias, retardos, faltas,
        vacaciones, incapacidad, permisos, comisiones, rotaciones,
        pct_asistencia, pct_retardo, pct_falta, pct_presencia
    """
    filas = []
    for id_emp in df["ID"].unique():
        df_emp  = df[df["ID"] == id_emp]
        primera = df_emp.iloc[0]
        metricas = _conteos_empleado(df_emp)

        filas.append({
            "ID":  id_emp,
            "Nombre_Completo":   str(primera.get("Nombre_Completo", "")),
            "Tipo":              str(primera.get("Tipo", "")),
            "Especialidad": str(primera.get("Especialidad", "")),
            "Grado":             str(primera.get("Grado", "")),
            "Periodo_Ingreso":   str(primera.get("Periodo_Ingreso", "")),
            **metricas,
        })

    return pd.DataFrame(filas).sort_values("Nombre_Completo").reset_index(drop=True)


def ranking_asistencia(df: pd.DataFrame,
                        top_n: int = 10,
                        ascendente: bool = False) -> pd.DataFrame:
    """
    Retorna los N empleados ordenados por % de asistencia puntual.

    ascendente=True  → los de peor asistencia primero (útil para reportes de riesgo).
    ascendente=False → los de mejor asistencia primero (reconocimiento).
    """
    stats = estadisticas_por_empleado(df)
    return (
        stats
        .sort_values("pct_asistencia", ascending=ascendente)
        .head(top_n)
        [["Nombre_Completo", "Tipo", "Especialidad", "Grado",
          "total_lab", "asistencias", "retardos", "faltas",
          "pct_asistencia", "pct_presencia"]]
        .reset_index(drop=True)
    )


# ──────────────────────────────────────────────
# Agregados por grupo
# ──────────────────────────────────────────────
def resumen_por_tipo(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega métricas por Tipo de personal (Interno, Residente, Médico Adscrito).
    Útil para comparativas entre grupos en el dashboard.
    """
    lab = df[df["Estatus"].isin(ESTATUSES_LABORABLES)]
    grupos = []
    for tipo, grp in lab.groupby("Tipo"):
        c = grp["Estatus"].value_counts().to_dict()
        total = len(grp)
        n_emp = grp["ID"].nunique()
        grupos.append({
            "Tipo":             tipo,
            "Empleados":        n_emp,
            "Días evaluados":   total,
            "Asistencias":      c.get("ASISTENCIA", 0),
            "Retardos":         c.get("RETARDO",    0),
            "Faltas":           c.get("FALTA",      0),
            "Incapacidades":    c.get("INCAPACIDAD",0),
            "Vacaciones":       c.get("VACACIONES", 0),
            "Rotaciones":       c.get("ROTACION",   0),
            "% Asistencia":     _safe_pct(c.get("ASISTENCIA", 0), total),
            "% Falta":          _safe_pct(c.get("FALTA",      0), total),
        })
    return pd.DataFrame(grupos)


def resumen_por_especialidad(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega métricas por Especialidad (excluyendo Internos que no tienen especialidad).
    """
    lab = df[
        df["Estatus"].isin(ESTATUSES_LABORABLES) &
        df["Especialidad"].str.strip().ne("")
    ]
    grupos = []
    for esp, grp in lab.groupby("Especialidad"):
        c = grp["Estatus"].value_counts().to_dict()
        total = len(grp)
        grupos.append({
            "Especialidad":   esp,
            "Empleados":      grp["ID"].nunique(),
            "Días evaluados": total,
            "Asistencias":    c.get("ASISTENCIA", 0),
            "Retardos":       c.get("RETARDO",    0),
            "Faltas":         c.get("FALTA",      0),
            "% Asistencia":   _safe_pct(c.get("ASISTENCIA", 0), total),
            "% Falta":        _safe_pct(c.get("FALTA",      0), total),
        })
    return pd.DataFrame(grupos).sort_values("% Falta", ascending=False)


# ──────────────────────────────────────────────
# Análisis temporal
# ──────────────────────────────────────────────
def dias_criticos(df: pd.DataFrame, top_n: int = 5) -> pd.DataFrame:
    """
    Identifica los días con mayor número de faltas + retardos en el período.
    Útil para detectar patrones (ej. lunes con más ausencias).

    Retorna DataFrame con: Fecha, DiaSemana, Faltas, Retardos, Total_Incidencias.
    """
    DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves",
               "Viernes", "Sábado", "Domingo"]

    lab = df[df["Estatus"].isin({"FALTA", "RETARDO"})].copy()
    lab["Fecha_dt"] = pd.to_datetime(lab["Fecha"])

    agg = (
        lab.groupby("Fecha_dt")["Estatus"]
        .value_counts()
        .unstack(fill_value=0)
        .reset_index()
    )

    # Asegurar columnas aunque no existan
    for col in ["FALTA", "RETARDO"]:
        if col not in agg.columns:
            agg[col] = 0

    agg["Total_Incidencias"] = agg.get("FALTA", 0) + agg.get("RETARDO", 0)
    agg["DiaSemana"] = agg["Fecha_dt"].dt.weekday.map(
        lambda x: DIAS_ES[x] if x < len(DIAS_ES) else str(x)
    )
    agg = agg.rename(columns={
        "Fecha_dt": "Fecha",
        "FALTA":    "Faltas",
        "RETARDO":  "Retardos",
    })

    return (
        agg[["Fecha", "DiaSemana", "Faltas", "Retardos", "Total_Incidencias"]]
        .sort_values("Total_Incidencias", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )


def calcular_tendencia_semanal(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrupa los días evaluados por semana del mes y calcula la distribución
    de estatus. Permite ver si la asistencia mejora o empeora con las semanas.

    Retorna DataFrame con columnas: Semana, Asistencias, Retardos, Faltas,
    Total_Laborables, Pct_Asistencia.
    """
    lab = df[df["Estatus"].isin(ESTATUSES_LABORABLES)].copy()
    lab["Fecha_dt"] = pd.to_datetime(lab["Fecha"])
    lab["Semana"]   = lab["Fecha_dt"].dt.isocalendar().week.astype(int)

    # Re-numerar semanas como Semana 1, 2, 3… desde la primera del mes
    semana_min = lab["Semana"].min()
    lab["Semana"] = lab["Semana"] - semana_min + 1

    filas = []
    for sem, grp in lab.groupby("Semana"):
        c = grp["Estatus"].value_counts().to_dict()
        total = len(grp)
        filas.append({
            "Semana":          f"Semana {int(sem)}",
            "Asistencias":     c.get("ASISTENCIA", 0),
            "Retardos":        c.get("RETARDO",    0),
            "Faltas":          c.get("FALTA",      0),
            "Total_Laborables": total,
            "Pct_Asistencia":  _safe_pct(c.get("ASISTENCIA", 0), total),
        })

    return pd.DataFrame(filas)


# ──────────────────────────────────────────────
# Exportación de estadísticas a Excel
# ──────────────────────────────────────────────
def exportar_estadisticas_excel(df: pd.DataFrame,
                                  directorio_salida: str,
                                  ruta_logo: Optional[str] = None) -> str:
    """
    Genera un archivo Excel con 4 hojas de estadísticas:
      1. Resumen_Individual  → métricas por empleado
      2. Por_Tipo            → agregado por tipo de personal
      3. Por_Especialidad    → agregado por especialidad
      4. Tendencia_Semanal   → evolución semana a semana
      5. Dias_Criticos       → días con más incidencias

    Retorna la ruta del archivo generado.
    """
    fechas  = pd.to_datetime(df["Fecha"])
    mes     = int(fechas.dt.month.mode()[0])
    anio    = int(fechas.dt.year.mode()[0])
    mes_str = MESES_ES.get(mes, str(mes))

    nombre_arch = f"SGAM_Estadisticas_{mes_str}{anio}.xlsx"
    Path(directorio_salida).mkdir(parents=True, exist_ok=True)
    ruta_out = Path(directorio_salida) / nombre_arch

    wb = openpyxl.Workbook()

    # ── Estilos ──────────────────────────────────────────────────────────
    FILL_H  = PatternFill("solid", fgColor="1F3864")
    FILL_A  = PatternFill("solid", fgColor="EBF3FB")
    FILL_Z  = PatternFill("solid", fgColor="F2F7FC")
    FNT_H   = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    FNT_T   = Font(name="Calibri", bold=True, size=13, color="1F3864")
    FNT_B   = Font(name="Calibri", size=9)
    FNT_G   = Font(name="Calibri", size=9, color="217346", bold=True)
    FNT_R   = Font(name="Calibri", size=9, color="C00000", bold=True)
    ALIN_C  = Alignment(horizontal="center", vertical="center")
    ALIN_L  = Alignment(horizontal="left",   vertical="center")
    BRD     = Border(*[Side(style="thin")] * 0,
                      left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"),  bottom=Side(style="thin"))

    def _hdr(ws, fila, valores, anchos):
        """Escribe una fila de encabezado y ajusta anchos."""
        for j, (val, ancho) in enumerate(zip(valores, anchos), start=1):
            c = ws.cell(row=fila, column=j, value=val)
            c.font = FNT_H; c.fill = FILL_H; c.alignment = ALIN_C; c.border = BRD
            ws.column_dimensions[get_column_letter(j)].width = ancho

    def _fila_data(ws, fila, valores, alterno=False):
        fill = FILL_Z if alterno else None
        for j, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=j, value=val)
            c.font = FNT_B
            if fill:
                c.fill = fill
            c.alignment = ALIN_L if j == 2 else ALIN_C
            c.border = BRD
            # Colorear porcentajes
            if isinstance(val, float):
                if "pct_falta" in str(ws.cell(row=1, column=j).value or "").lower():
                    c.font = FNT_R if val > 20 else FNT_B
                elif "pct_asistencia" in str(ws.cell(row=1, column=j).value or "").lower():
                    c.font = FNT_G if val >= 80 else FNT_B

    def _titulo_hoja(ws, texto):
        ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A1:{get_column_letter(ws.max_column or 15)}1")
        c = ws.cell(row=1, column=1, value=texto)
        c.font = FNT_T; c.fill = FILL_A
        c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Hoja 1: Resumen Individual ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen_Individual"
    ws1.row_dimensions[2].height = 18

    df_ind = estadisticas_por_empleado(df)
    cols_h = ["#", "Nombre", "Tipo", "Especialidad", "Grado", "Periodo",
              "Días Lab.", "Asist.", "Retardos", "Faltas", "Vacac.",
              "Incap.", "Permisos", "Comis.", "Rotac.",
              "% Asist.", "% Retardo", "% Falta", "% Presencia"]
    anchos = [4, 28, 14, 16, 6, 7, 8, 7, 7, 7, 7, 7, 7, 7, 7, 9, 9, 9, 10]

    _hdr(ws1, 2, cols_h, anchos)

    for idx, row in df_ind.iterrows():
        fila = idx + 3
        ws1.row_dimensions[fila].height = 14
        vals = [
            idx + 1,
            row["Nombre_Completo"],   row["Tipo"],
            row["Especialidad"], row["Grado"],
            row["Periodo_Ingreso"],   row["total_lab"],
            row["asistencias"],       row["retardos"],
            row["faltas"],            row["vacaciones"],
            row["incapacidad"],       row["permisos"],
            row["comisiones"],        row["rotaciones"],
            row["pct_asistencia"],    row["pct_retardo"],
            row["pct_falta"],         row["pct_presencia"],
        ]
        _fila_data(ws1, fila, vals, alterno=idx % 2 == 0)

    _titulo_hoja(ws1, f"SGAM – Estadísticas Individuales  |  {mes_str} {anio}")
    ws1.freeze_panes = "A3"

    # ── Hoja 2: Por Tipo ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por_Tipo")
    ws2.row_dimensions[2].height = 18
    df_tipo = resumen_por_tipo(df)
    if not df_tipo.empty:
        cols_t = list(df_tipo.columns)
        anchos_t = [20, 10, 14, 12, 10, 10, 12, 10, 12, 13, 10]
        _hdr(ws2, 2, cols_t, anchos_t[:len(cols_t)])
        for idx, row in df_tipo.iterrows():
            fila = idx + 3
            ws2.row_dimensions[fila].height = 14
            _fila_data(ws2, fila, list(row), alterno=idx % 2 == 0)
    _titulo_hoja(ws2, f"SGAM – Resumen por Tipo de Personal  |  {mes_str} {anio}")

    # ── Hoja 3: Por Especialidad ─────────────────────────────────────────
    ws3 = wb.create_sheet("Por_Especialidad")
    ws3.row_dimensions[2].height = 18
    df_esp = resumen_por_especialidad(df)
    if not df_esp.empty:
        cols_e = list(df_esp.columns)
        anchos_e = [22, 10, 14, 12, 10, 10, 13, 10]
        _hdr(ws3, 2, cols_e, anchos_e[:len(cols_e)])
        for idx, row in df_esp.iterrows():
            fila = idx + 3
            ws3.row_dimensions[fila].height = 14
            _fila_data(ws3, fila, list(row), alterno=idx % 2 == 0)
    _titulo_hoja(ws3, f"SGAM – Resumen por Especialidad  |  {mes_str} {anio}")

    # ── Hoja 4: Tendencia Semanal ────────────────────────────────────────
    ws4 = wb.create_sheet("Tendencia_Semanal")
    ws4.row_dimensions[2].height = 18
    df_sem = calcular_tendencia_semanal(df)
    if not df_sem.empty:
        cols_s = list(df_sem.columns)
        anchos_s = [14, 12, 10, 10, 16, 16]
        _hdr(ws4, 2, cols_s, anchos_s[:len(cols_s)])
        for idx, row in df_sem.iterrows():
            fila = idx + 3
            ws4.row_dimensions[fila].height = 14
            _fila_data(ws4, fila, list(row), alterno=idx % 2 == 0)
    _titulo_hoja(ws4, f"SGAM – Tendencia Semanal  |  {mes_str} {anio}")

    # ── Hoja 5: Días Críticos ────────────────────────────────────────────
    ws5 = wb.create_sheet("Dias_Criticos")
    ws5.row_dimensions[2].height = 18
    df_crit = dias_criticos(df, top_n=15)
    if not df_crit.empty:
        cols_c = list(df_crit.columns)
        anchos_c = [14, 14, 10, 10, 18]
        _hdr(ws5, 2, cols_c, anchos_c[:len(cols_c)])
        for idx, row in df_crit.iterrows():
            fila = idx + 3
            ws5.row_dimensions[fila].height = 14
            vals = [
                str(row["Fecha"].date()) if hasattr(row["Fecha"], "date")
                else str(row["Fecha"]),
                row["DiaSemana"],
                row["Faltas"],
                row["Retardos"],
                row["Total_Incidencias"],
            ]
            _fila_data(ws5, fila, vals, alterno=idx % 2 == 0)
    _titulo_hoja(ws5, f"SGAM – Días con Más Incidencias  |  {mes_str} {anio}")

    wb.save(str(ruta_out))
    return str(ruta_out)
